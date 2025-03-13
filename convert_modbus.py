#! /usr/bin/env python3
# -*- coding: utf-8 -*-
# vim:fenc=utf-8
#
# Copyright © 2025 jorgen <jorgen@jorgen-pc>
#
# Distributed under terms of the MIT license.

import re
import json
from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

wb = load_workbook('./assets/modbus-registers.xlsx')

ws = wb['Field list']

print('[');

first = True
for row in ws.iter_rows(min_row=3):
    if row[6].value == '':
        continue

    if first:
        first = False
    else:
        print(',')

    range = [ float(v.strip()) for v in (row[5].value or '').split('to') if re.match(r'^[-+]?\d+(.\d+)?$', v.strip()) ]
    if len(range) < 2: range = None

    print(json.dumps({
        'service': row[0].value.split('.')[-1],
        'name': row[1].value,
        'type': row[3].value,
        'range': range,
        'scalefactor': row[4].value,
        'topic': row[6].value,
        'writable': row[7].value == 'yes',
        'unit': row[8].value
    }), end='')

print( ']' );
