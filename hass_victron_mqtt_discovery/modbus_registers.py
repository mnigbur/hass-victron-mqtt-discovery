#! /usr/bin/env python3
# -*- coding: utf-8 -*-
# vim:fenc=utf-8
#
# Copyright © 2025 jorgen <jorgen@jorgen-pc>
#
# Distributed under terms of the MIT license.
import re
import json

from openpyxl import Workbook
from openpyxl.reader.excel import load_workbook

from .cache import Cache

def topic_from_row(row):
    if row[6] == '':
        return None

    value_range = [ float(v.strip()) for v in (row[5] or '').split('to') if re.match(r'^[-+]?\d+(.\d+)?$', v.strip()) ]

    if len(value_range) < 2:
        value_range = None

    return {
        'service': row[0].split('.')[-1],
        'name': row[1],
        'type': row[3],
        'range': value_range,
        'scalefactor': row[4],
        'topic': row[6],
        'writable': row[7] == 'yes',
        'unit': row[8]
    }


class ModbusRegisters(Cache):

    def __init__(self, source=None):
        self.services = {}

        if source is not None:
            self.from_xlsx(source)

    def lookup_mqtt_topic(self, topic):
        m = re.search(r'N/([^/]+)/([^/]+)/(\d+)(/.*)$', topic)
        if m is None:
            return None

        service = m.group(2).lower()
        dbusTopic = m.group(4)

        if not service in self.services:
            return None

        topicMap = self.services[service]
        searchTopic = dbusTopic.lower()
        if not searchTopic in topicMap: searchTopic = re.sub(r'/power$', '/p', searchTopic)
        if not searchTopic in topicMap: searchTopic = re.sub(r'/current$', '/i', searchTopic)
        if not searchTopic in topicMap: searchTopic = re.sub(r'/voltage$', '/v', searchTopic)
        if not searchTopic in topicMap: searchTopic = re.sub(r'/frequency$', '/f', searchTopic)

        if not searchTopic in topicMap:
            return None

        return topicMap[searchTopic]

    def from_xlsx(self, file):
        cache_id = self.file_cache_id(file)

        if self.has_cache(cache_id):
            self.services = self.load_cache(cache_id)
            return

        wb = load_workbook(file)
        ws = wb['Field list']

        for row in ws.iter_rows(min_row=3):
            row = [ cell.value for cell in row ]
            row = topic_from_row(row)

            if row is None:
                continue

            service = row['service']
            topic = row['topic']

            if service is None or topic is None:
                 continue

            service = service.lower()
            topic = topic.lower()

            if service not in self.services:
                 self.services[service] = {}

            if not topic in self.services[service]:
                self.services[service][topic] = row

        wb.close()
        self.update_cache(cache_id, self.services)
